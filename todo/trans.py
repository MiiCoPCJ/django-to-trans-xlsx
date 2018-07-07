import os
import re
from openpyxl import *
from common.format import *

# 执行文件的路径
path = os.path.split(os.path.realpath(__file__))[0]
# 读取model数据 ,去除#注明部分

def make_xlsx(pre,file_path,wb,i):
    str = ''

    if i>0:
        sheet = wb.create_sheet(title=pre)
    else:
        sheet = wb.active
        sheet.title = pre

    # sheet的行r和列c
    r = 1
    c = 1

    f = open(file_path,'r',encoding='utf-8')
    for line in f.readlines():
        line = line.strip()
        if not len(line) or line.startswith('#'):
            continue
        str += line + '\n'
    f.close()

    # 去除"""注释
    str = re.sub('"""(.|\n)*?"""','',str)

    # 正则 查找
    pattern = re.compile(r'class\s+\w+\((AbstractBaseUser|models.Model|)\):.*?verbose_name\s?=\s?verbose_name_plural\s?=\s?(\'|")(\w|-)+(\'|")', re.S)

    find = re.finditer(pattern,str)

    # 历遍
    for m in find:
        model = m.group()
        print(model)
        print('####')

        # 表名 英文
        data_name = re.search('class\s+(\w+)\(.*?\):', model).group(1).lower()
        # 表名 中文
        data_name_cn = re.search('verbose_name_plural\s?=\s?(\'|")((\w|-)+)(\'|")', model).group(2)

        sheet.cell(row=r,column=c,value='表名称:')
        sheet.cell(row=r, column=c+1, value=data_name_cn+'('+ pre+'_' + data_name+')')
        r = r + 2
        line_title(sheet,r)
        r = r + 1

        # 获取字段部分
        field = re.search('class\s+\w+\(.*?\):(.*)class\sMeta:', model, re.S).group(1)
        stp = field.strip().split('\n')
        length = len(stp)
        # 外键个数
        k = len(re.findall('ForeignKey',field))
        Tip = True
        # 添加外键计算
        y = 0
        # 添加字段计算
        b = 0
        # 选择数据说明计算
        u = 0

        for sp in stp:
            # 选择数据说明
            if not re.findall('models.',sp):
                if re.findall(',',sp):
                    sheet.cell(row=r + u, column=7, value=sp)
                else:
                    sheet.cell(row=r + u, column=6, value=sp)
                u = u + 1

        for sp in stp:
            # 外键处理
            key = re.search('(\w+)\s?=\s?models\.ForeignKey\((.*)\)',sp)
            if key is not None and len(key.group()):
                key_name = key.group(1)
                foreign = key.group(2).split(',')[0].strip()
                # 查找外键所在表
                link_key = re.search('from\s(\w+).models\simport\s'+foreign, str)
                if link_key is not None and len(link_key.group()):
                    dl = link_key.group(1) + '_' + foreign.lower()
                else:
                    dl = pre + '_' + foreign.lower()

                if Tip:
                    foreign_title(sheet,r,length - u)
                    Tip = False
                sheet.cell(row=r + length - u + 2 + y, column=1, value=key_name)
                sheet.cell(row=r + length - u + 2 + y, column=2, value=dl)

                des = re.search('\s?'+foreign+'\s?,(.*)',key.group(2)).group(1).strip()
                sheet.cell(row=r + length - u + 2 + y, column=3, value=des)

                y = y + 1

                continue



            d = re.search('(\w+)\s?=\s?models\.(\w+)Field\((.*)\)',sp)

            if d is not None and len(d.group()):


                field_name = d.group(1)
                field_type = d.group(2)
                if field_type == 'DateTime':
                    sheet.cell(row=r + b, column=1, value=d.group(3))
                    sheet.cell(row=r + b, column=2, value=field_name)
                    continue

                if re.search('verbose_name\s?=\s?(\'|")((\w|-)+)(\'|")', d.group(3)):
                    field_name_cn = re.search('verbose_name\s?=\s?(\'|")((\w|-)+)(\'|")', d.group(3)).group(2)
                else:
                    field_name_cn = re.sub('"','',d.group(3).split(',')[0])

                sheet.cell(row=r + b, column=1, value=field_name_cn)
                sheet.cell(row=r + b, column=2, value=field_name)

                if field_type == 'Decimal':
                    max_digits = re.search('max_digits\s?=\s?(\d+)\s?(,)?', d.group(3)).group(1)
                    decimal_places = re.search('decimal_places\s?=\s?(\d+)\s?(,)?', d.group(3)).group(1)
                    sheet.cell(row=r + b, column=3, value=field_type+'('+max_digits+','+decimal_places+')')
                elif field_type == 'Char':
                    max_length = re.search('max_length\s?=\s?(\d+)\s?(,)?', d.group(3)).group(1)
                    sheet.cell(row=r + b, column=3, value=field_type + '(' + max_length + ')')
                else:
                    sheet.cell(row=r + b, column=3, value=field_type)

                # 默认值
                if re.search('default\s?=\s?(.*)(,)?',d.group(3)):
                    dft = re.search('default\s?=\s?(.*)(,)?',d.group(3)).group(1)
                    sheet.cell(row=r + b, column=4, value=dft)

                # 描述
                for dt in d.group(3).split(','):
                    de = re.search('(\w+)=.*',dt)
                    if de:
                        if de.group(1) != 'verbose_name' and de.group(1)!='max_digits' and de.group(1)!='max_length' and de.group(1)!='default'\
                                and de.group(1) != 'decimal_places':
                            sheet.cell(row=r + b, column=5, value=dt)

                b = b + 1

        if (b+k+6)<(u+1):
            r = r + u + 1
        else:
            r = r + b + k + 6


    wb.save(path + '/../doc/model.xlsx')





# 查找resource目录所有model文件
def loop_res():
    dir = path + "/../resource/"

    # 开启
    wb = Workbook()
    i = 0
    for parent, dirnames, filenames in os.walk(dir):
        for filename in filenames:
            file_path = os.path.join(parent, filename)

            pre = re.sub('_models.py','',filename)
            make_xlsx(pre,file_path,wb,i)
            i =i + 1

if __name__ == "__main__":
     loop_res()
    # wb = Workbook()
    # make_xlsx('config', path+'/../resource/wc_auth_models.py', wb, 0)